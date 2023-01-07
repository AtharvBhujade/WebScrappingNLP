from openpyxl import Workbook, load_workbook
from textProccessor import *

wb = load_workbook("Output Data Structure.xlsx")
ws = wb.active
# print(ws)

for i in range(1, 171):
    filename = "Final\\" + str(i) + ".txt"
    text = readText(filename)
    cleanText = getCleanText(text)
    tokenizedText = getTokenizedText(cleanText)
    finalText = removeStopword(tokenizedText)
    finalNonTokenText = ' '.join(finalText)
    pos = getPositiveScore(finalNonTokenText)
    neg = getNegativeScore(finalNonTokenText)
    cellNumber = i + 1

    cellAlpha = "C"
    cellCord = cellAlpha + str(cellNumber)
    ws[cellCord].value = float(getPositiveScore(finalNonTokenText))

    cellAlpha = "D"
    cellCord = cellAlpha + str(cellNumber)
    ws[cellCord].value = float(getNegativeScore(finalNonTokenText))

    cellAlpha = "E"
    cellCord = cellAlpha + str(cellNumber)
    ws[cellCord].value = float(getPolarityScore(positiveScore=getPositiveScore(finalNonTokenText),
                                                negativeScore=getNegativeScore(finalNonTokenText)))

    cellAlpha = "F"
    cellCord = cellAlpha + str(cellNumber)
    ws[cellCord].value = float(getSubjectivityScore(pos, neg, tokenizedText))

    cellAlpha = "G"
    cellCord = cellAlpha + str(cellNumber)
    ws[cellCord].value = float(getAverageSentencelength(text))

    cellAlpha = "J"
    cellCord = cellAlpha + str(cellNumber)
    ws[cellCord].value = float(getAverageNumberofWordPerSentence(text, tokenizedText))

    cellAlpha = "L"
    cellCord = cellAlpha + str(cellNumber)
    ws[cellCord].value = float(getWordCount(tokenizedText))

    cellAlpha = "O"
    cellCord = cellAlpha + str(cellNumber)
    ws[cellCord].value = float(getAverageWordLength(tokenizedText))

wb.save("Output Data Structure.xlsx")